# -*- coding: utf-8 -*-
"""
Created on Mon Nov 22 15:03:38 2021

@author: zaloginv
"""

import openpyxl as opx
from openpyxl.styles import Alignment
import time
import os
import datetime
import re

def follow(thefile):
    # поиск конца файла
    thefile.seek(0, os.SEEK_END)
    
    # запуск бесконечного цикла
    while True:
        
        try: # для прерывания цикла вручную, чтоб можно было сделать что-то с файлом
        
            # читаем последнюю строчку файла
            line = thefile.readline()
            # sleep если файл не обновляется
            if not line:
                time.sleep(0.1)
                continue
    
            yield line
            
        except KeyboardInterrupt:
            break


# просмотр url адресов
def url_seeker():
    path = '//KOMPUTER\site\infobeast'
    url_list = []
    # путь к сайту на сервере
    
    for root, dirs, files in os.walk(path):
        for file in files:
            url = os.path.join(root,file) # пишем полный путь
            url_list.append((url.replace(path,'')).replace('\\','/')) # отрезаем кусок пути
    url_list.extend(['/','', '/favicon.ico', '/apple-touch-icon.png'])
            
    return url_list



class base(): # базовые значения

    ips = {} # словарь айпишников
    
    logfile = open("//KOMPUTER\logs\mylog.log","r")
    loglines = follow(logfile)
    
    site = url_seeker()
    
    tab = '\t' # разделитель
    dateformat = '%d/%b/%Y %H:%M:%S.%f' # формат даты
    
    SQLi = [
        'SELECT', 'AND', '=', "'"
        ]
    
    pars = [
        'nw'
        ]
    
    locked_urls = [
        '.htaccess', 'httpd.conf'
        ]
    
    ban_symb = 'V' # обозначение бан пользователя
    unban_symb = 'O' # обозначение разбан пользователя
    
    # доступ к файлу, по которому банятся нарушители
    htaccess_adress = '//KOMPUTER/site/infobeast/.htaccess'
    original = ['Require all granted'] # первые строки для бана
    
    report_loc = 'report/report.xlsx' # куда кладется отчет
    
        ############# счётчики ############
    xl_All = 1
    
    xl_SQL = 1
    
    xl_Sus = 1
    
    xl_DOS = 1 # счётчик для строк эксель
    DOS_timing = 1 # таймер времени (если прошло меньше указанного времени, то...)
    max_DOS = 1000 # максимум атак для бана по DOS
    min_DOS = 999 # минимум для начала отсчёта атак по DOS

    xl_Pars = 1
    Pars_timing = 20 # таймер времени
    max_Pars = 2 # максимум совпадений по парсингу
    
    xl_Loc = 1
    
    
    

def banning(ip): # бан пользователя
    htaccess = open(base.htaccess_adress, 'a') # открываем ещё раз на запись
    htaccess.writelines(f'\nRequire not ip {ip}')
    htaccess.close()
    
    

def main():    
    
    ############## формирование таблицы Excel #####################  
    
    wb = opx.Workbook() #создаём рабочую книгу
    wsAll = wb.active #при создании рабочей книги, появляется первый лист - активный, присваиваем ему имя
    wsAll.title = ("Обработанные IP")
    wsDOS = wb.create_sheet("DOS-атаки")
    wsPars = wb.create_sheet("Парсинг")
    wsSus = wb.create_sheet("Подозрительная активность")
    wsSQL = wb.create_sheet('SQL-инъекции')
    wsLoc = wb.create_sheet('Запрещенные ссылки')
    
    ############## All ##############
    
    wsAll['A1'] = ( 'IP-адрес' )
    wsAll['A1'].alignment = Alignment(horizontal='center')
    
    wsAll['B1'] = ( 'Дата и время' )
    wsAll['B1'].alignment = Alignment(horizontal='center')
    
    wsAll['C1'] = ( 'Содержимое строки' )
    wsAll['C1'].alignment = Alignment(horizontal='center')
    
    wsAll['D1'] = ( 'Код доступа' )
    wsAll['D1'].alignment = Alignment(horizontal='center')
    
    wsAll['E1'] = ( f'Статус доступа ({base.ban_symb} - забанен, {base.unban_symb} - есть доступ)' )
    wsAll['E1'].alignment = Alignment(horizontal='center')
    
    
    ############## SQLi ##############
    
    wsSQL['A1'] = ( 'IP-адрес' )
    wsSQL['A1'].alignment = Alignment(horizontal='center')
    
    wsSQL['B1'] = ( 'Дата и время' )
    wsSQL['B1'].alignment = Alignment(horizontal='center')
    
    wsSQL['C1'] = ( 'Содержимое строки' )
    wsSQL['C1'].alignment = Alignment(horizontal='center')
    
    wsSQL['D1'] = ( 'Код доступа' )
    wsSQL['D1'].alignment = Alignment(horizontal='center')
    
    wsSQL['E1'] = ( f'Статус доступа ({base.ban_symb} - забанен, {base.unban_symb} - есть доступ)' )
    wsSQL['E1'].alignment = Alignment(horizontal='center')


    ############## DOS ##############
    
    wsDOS['A1'] = ( 'IP-адрес' )
    wsDOS['A1'].alignment = Alignment(horizontal='center')
    
    wsDOS['B1'] = ( 'Дата и время' )
    wsDOS['B1'].alignment = Alignment(horizontal='center')
    
    wsDOS['C1'] = ( 'Суммарное количество атак' )
    wsDOS['C1'].alignment = Alignment(horizontal='center')
    
    wsDOS['D1'] = ( 'Код доступа' )
    wsDOS['D1'].alignment = Alignment(horizontal='center')
    
    wsDOS['E1'] = ( f'Статус доступа ({base.ban_symb} - забанен, {base.unban_symb} - есть доступ)' )
    wsDOS['E1'].alignment = Alignment(horizontal='center')
    
    ############## Parsing ##############
    
    wsPars['A1'] = ( 'IP-адрес' )
    wsPars['A1'].alignment = Alignment(horizontal='center')
    
    wsPars['B1'] = ( 'Дата и время' )
    wsPars['B1'].alignment = Alignment(horizontal='center')
    
    wsPars['C1'] = ( 'Содержимое строки' )
    wsPars['C1'].alignment = Alignment(horizontal='center')

    ############## подозрительная активность ##############
    
    wsSus['A1'] = ( 'IP-адрес' )
    wsSus['A1'].alignment = Alignment(horizontal='center')
    
    wsSus['B1'] = ( 'Дата и время' )
    wsSus['B1'].alignment = Alignment(horizontal='center')
    
    wsSus['C1'] = ( 'Содержимое строки' )
    wsSus['C1'].alignment = Alignment(horizontal='center')
    
    wsAll['D1'] = ( 'Код доступа' )
    wsAll['D1'].alignment = Alignment(horizontal='center')
    
    ############# запрещенные ссылки ###############

    wsLoc['A1'] = ( 'IP-адрес' )
    wsLoc['A1'].alignment = Alignment(horizontal='center')
    
    wsLoc['B1'] = ( 'Дата и время' )
    wsLoc['B1'].alignment = Alignment(horizontal='center')
    
    wsLoc['C1'] = ( 'Содержимое строки' )
    wsLoc['C1'].alignment = Alignment(horizontal='center')
    
    wsLoc['D1'] = ( 'Код доступа' )
    wsLoc['D1'].alignment = Alignment(horizontal='center')
    
    wsLoc['E1'] = ( f'Статус доступа ({base.ban_symb} - забанен, {base.unban_symb} - есть доступ)' )
    wsLoc['E1'].alignment = Alignment(horizontal='center')    
    
    

    ############## обработка строк ############### 
    
    ''' описание некоторых переменных
    req - количество запросов всего
    time - время, в которое пользователь что-то сделал
    url - запрашиваемая строка url
    dos - счётчик возможной dos-атаки
    '''
    
    
    
    #with open (htaccess_adress, 'r') as htaccess:
        #original = htaccess.readlines() # считываем все строки в режиме чтения
    
    htaccess = open(base.htaccess_adress, 'w') # открываем на запись
    htaccess.writelines(base.original)
    htaccess.close()
    

    
    # итерация через генератор
    for line in base.loglines:
        
        ip = str(line.split(base.tab)[2]).split(':')[3]
        
        # если такого ip ещё нет в словаре
        if ip not in base.ips:
            base.ips.update({ip:{'req':0, 'time':None, 'url':'', 'parsing':0, 'code':'', 'status':''}}) # добавляем ip как ключ 
            base.ips[ip]['req'] = 1
            base.ips[ip]['dos'] = 0
            base.ips[ip]['time'] = datetime.datetime.strptime((line.split(base.tab)[0]), base.dateformat)
            base.ips[ip]['url'] = str(line.split(base.tab)[3]).split(' ')[1]
            base.ips[ip]['parsing'] = 0
            base.ips[ip]['code'] = str(line.split(base.tab)[4])
            base.ips[ip]['status'] = base.unban_symb
            newtime = datetime.datetime.strptime((line.split(base.tab)[0]), base.dateformat)
            
            
        # если такой ip есть в словаре
        elif ip in base.ips:
            base.ips[ip]['req'] += 1
            newtime = datetime.datetime.strptime((line.split(base.tab)[0]), base.dateformat)
            if abs((base.ips[ip]['time'] - newtime)) < datetime.timedelta(seconds = base.DOS_timing): # сравниваем, сколько секунд прошло с предыдущего запроса
                   base.ips[ip]['dos'] += 1 # если прошло меньше 2, то счётчик дос-атак увеличиается
                   
        # dos-атака
                   if base.ips[ip]['dos'] > base.min_DOS and base.ips[ip]['dos'] < base.max_DOS:
                       base.xl_DOS += 1
                       wsDOS[f'A{base.xl_DOS}']=ip # IP-адрес
                       wsDOS[f'B{base.xl_DOS}']=base.ips[ip]['time'] # дата и время
                       wsDOS[f'C{base.xl_DOS}']=base.ips[ip]['dos'] # суммарное количество атак
                       wsDOS[f'D{base.xl_DOS}']=base.ips[ip]['code'] # код доступа
                       wsDOS[f'E{base.xl_DOS}']=base.ips[ip]['status'] # статус бана
                
                   elif base.ips[ip]['dos'] >= base.max_DOS:
                       base.xl_DOS += 1
                       wsDOS[f'A{base.xl_DOS}']=ip # IP-адрес
                       wsDOS[f'B{base.xl_DOS}']=base.ips[ip]['time'] # дата и время
                       wsDOS[f'C{base.xl_DOS}']=base.ips[ip]['dos'] # суммарное количество атак
                       wsDOS[f'D{base.xl_DOS}']=base.ips[ip]['code'] # код доступа
                       if base.ips[ip]['status'] == base.unban_symb:
                           base.ips[ip]['status'] = base.ban_symb
                           banning(ip) # бан
                       wsDOS[f'E{base.xl_DOS}']=base.ips[ip]['status'] # статус
                       

        

        if str(base.ips[ip]['url']) not in base.site:

        # SQL-атака            
            for symbol in base.SQLi:
                if str(symbol).lower() in str(base.ips[ip]['url']):
                    base.xl_SQL += 1
                    wsSQL[f'A{base.xl_SQL}']=ip # IP-адрес
                    wsSQL[f'B{base.xl_SQL}']=base.ips[ip]['time'] # дата и время
                    wsSQL[f'C{base.xl_SQL}']=base.ips[ip]['url'] # строка
                    wsSQL[f'D{base.xl_SQL}']=base.ips[ip]['code'] # код доступа
                    if base.ips[ip]['status'] == base.unban_symb:
                        base.ips[ip]['status'] = base.ban_symb
                        banning(ip) # бан
                    wsSQL[f'E{base.xl_SQL}']=base.ips[ip]['status'] # статус
                    break
        
        # не sqli, но странные url        
                else:
                    base.xl_Sus += 1
                    wsSus[f'A{base.xl_Sus}']=ip # IP-адрес
                    wsSus[f'B{base.xl_Sus}']=base.ips[ip]['time'] # дата и время
                    wsSus[f'C{base.xl_Sus}']=base.ips[ip]['url'] # строка
                    wsSus[f'D{base.xl_Sus}']=base.ips[ip]['code'] # код доступа
                    break
            
        # проверка на парсинг
        if str(base.ips[ip]['url']) in base.site:
            for page in base.pars: # для страниц, которые могут быть использованы в парсинге
                if page in str(base.ips[ip]['url']):
                    nums = re.findall(r'\d+', str(base.ips[ip]['url'])) # поиск всех чисел в строке
                    
                    try:
                        # nums[0] - поскольку у нас лишь одно число в строке url
                        if int(nums[0]) == 1:
                            base.ips[ip]['parsing'] = 0
                        
                        
                        elif ((int(nums[0]) - base.ips[ip]['parsing']) >= 1) and (abs((base.ips[ip]['time'] - newtime)) < datetime.timedelta(seconds = base.Pars_timing)): 
                        # если разница между запрашиваемой страничкой и предыдущей больше или равна 1   И   между запросами прошло меньше указанного времени
                            base.ips[ip]['parsing'] += 1
                            
                            if base.ips[ip]['parsing'] > base.max_Pars:
                                
                                base.xl_Pars += 1
                                wsPars[f'A{base.xl_Pars}']=ip # IP-адрес
                                wsPars[f'B{base.xl_Pars}']=base.ips[ip]['time'] # дата и время
                                wsPars[f'C{base.xl_Pars}']=base.ips[ip]['url'] # строка
                                
                                
                    except IndexError:
                        continue
                        
        
        # проверка на запрещенные ссылки
        for locked_url in base.locked_urls:
            if locked_url in str(base.ips[ip]['url']):
                base.xl_Loc += 1
                wsLoc[f'A{base.xl_Loc}']=ip # IP-адрес
                wsLoc[f'B{base.xl_Loc}']=base.ips[ip]['time'] # дата и время
                wsLoc[f'C{base.xl_Loc}']=base.ips[ip]['url'] # строка
                wsLoc[f'D{base.xl_Loc}']=base.ips[ip]['code'] # код доступа
                if base.ips[ip]['status'] == base.unban_symb:
                    base.ips[ip]['status'] = base.ban_symb
                    banning(ip) # бан
                wsLoc[f'E{base.xl_Loc}']=base.ips[ip]['status'] # статус                
        
        
                    
        # все обработанные строки
        base.xl_All += 1
        wsAll[f'A{base.xl_All}']=ip # IP-адрес
        wsAll[f'B{base.xl_All}']=base.ips[ip]['time'] # дата и время
        wsAll[f'C{base.xl_All}']=base.ips[ip]['url'] # содержимое строки
        wsAll[f'D{base.xl_All}']=base.ips[ip]['code'] # код доступа
        wsAll[f'E{base.xl_All}']=base.ips[ip]['status'] # статус бана
                    
        wb.save(base.report_loc)
                    
        base.ips[ip]['time'] = newtime # присваивается новое время
        base.ips[ip]['url'] = str(line.split(base.tab)[3]).split(' ')[1] # присваивается новый url
        base.ips[ip]['code'] = str(line.split(base.tab)[4])
                    
                

if __name__ == '__main__':
    main();

        
        
        
        
        
        
        
        